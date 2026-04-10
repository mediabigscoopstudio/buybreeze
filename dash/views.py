from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Count, Avg, Q
from django.utils import timezone
from datetime import timedelta
import io
import csv
import openpyxl
from django.http import HttpResponse

from .models import (
    Branch, UserProfile, Lead, CallLog, CallWrapUp, FollowUp, SystemSetting
)


# ============================================================
# AUTH GUARD
# ============================================================
def superadmin_required(user):
    return user.is_authenticated and user.is_superuser


# ============================================================
# AUTH VIEWS
# ============================================================
def login_view(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('http://localhost:8000/')
        if hasattr(request.user, 'profile'):
            role = request.user.profile.role
            if role == 'hr':      return redirect('http://hrpanel.localhost:8000/')
            if role == 'manager': return redirect('http://manager.localhost:8000/')
            if role == 'tl':      return redirect('http://tl.localhost:8000/')
        return redirect('http://localhost:8000/')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            if user.is_superuser:
                return redirect('http://localhost:8000/')
            if hasattr(user, 'profile'):
                role = user.profile.role
                if role == 'hr':      return redirect('http://hrpanel.localhost:8000/')
                if role == 'manager': return redirect('http://manager.localhost:8000/')
                if role == 'tl':      return redirect('http://tl.localhost:8000/')
            return redirect('http://localhost:8000/')
        else:
            messages.error(request, 'Invalid credentials.')

    return render(request, 'dash/signin.html')


def logout_view(request):
    logout(request)
    return redirect('http://localhost:8000/login/')


# ============================================================
# DASHBOARD HOME
# ============================================================
@user_passes_test(superadmin_required, login_url='/login/')
def index(request):
    total_leads       = Lead.objects.count()
    hot_leads         = Lead.objects.filter(temperature='hot').count()
    total_calls       = CallLog.objects.count()
    total_branches    = Branch.objects.filter(status='Enabled').count()
    total_users       = UserProfile.objects.filter(status='Enabled').count()
    pending_followups = FollowUp.objects.filter(followup_status='pending').count()

    stage_data   = Lead.objects.values('stage').annotate(count=Count('id'))
    branch_leads = Branch.objects.annotate(lead_count=Count('lead')).filter(status='Enabled')
    recent_leads = Lead.objects.order_by('-created_at')[:8]

    upcoming_followups = FollowUp.objects.filter(
        followup_status='pending',
        followup_at__gte=timezone.now(),
        followup_at__lte=timezone.now() + timedelta(hours=24)
    ).order_by('followup_at')[:5]

    context = {
        'total_leads': total_leads,
        'hot_leads': hot_leads,
        'total_calls': total_calls,
        'total_branches': total_branches,
        'total_users': total_users,
        'pending_followups': pending_followups,
        'stage_data': stage_data,
        'branch_leads': branch_leads,
        'recent_leads': recent_leads,
        'upcoming_followups': upcoming_followups,
    }
    return render(request, 'dash/index.html', context)


# ============================================================
# BRANCH MANAGEMENT
# ============================================================
@user_passes_test(superadmin_required, login_url='/login/')
def branch(request):
    branches = Branch.objects.all().order_by('-id')
    return render(request, 'dash/branch/branches.html', {'branches': branches})


@user_passes_test(superadmin_required, login_url='/login/')
def add_branch(request):
    if request.method == 'POST':
        Branch.objects.create(
            name       = request.POST.get('name'),
            location   = request.POST.get('location'),
            address    = request.POST.get('address'),
            phone      = request.POST.get('phone'),
            email      = request.POST.get('email'),
            gps_lat    = request.POST.get('gps_lat') or None,
            gps_lng    = request.POST.get('gps_lng') or None,
            gps_radius = request.POST.get('gps_radius') or 100,
        )
        messages.success(request, 'Branch added successfully.')
        return redirect('branch')
    return render(request, 'dash/branch/add_branch.html')


@user_passes_test(superadmin_required, login_url='/login/')
def edit_branch(request, id):
    item = get_object_or_404(Branch, id=id)
    if request.method == 'POST':
        item.name       = request.POST.get('name')
        item.location   = request.POST.get('location')
        item.address    = request.POST.get('address')
        item.phone      = request.POST.get('phone')
        item.email      = request.POST.get('email')
        item.gps_lat    = request.POST.get('gps_lat') or None
        item.gps_lng    = request.POST.get('gps_lng') or None
        item.gps_radius = request.POST.get('gps_radius') or 100
        item.save()
        messages.success(request, 'Branch updated successfully.')
        return redirect('branch')
    return render(request, 'dash/branch/edit_branch.html', {'data': item})


@user_passes_test(superadmin_required, login_url='/login/')
def delete_branch(request, id):
    item = get_object_or_404(Branch, id=id)
    item.delete()
    messages.success(request, 'Branch deleted.')
    return redirect('branch')


@user_passes_test(superadmin_required, login_url='/login/')
def enable_branch(request, id):
    item = get_object_or_404(Branch, id=id)
    item.status = 'Enabled'
    item.save()
    return redirect('branch')


@user_passes_test(superadmin_required, login_url='/login/')
def disable_branch(request, id):
    item = get_object_or_404(Branch, id=id)
    item.status = 'Disabled'
    item.save()
    return redirect('branch')


# ============================================================
# USER MANAGEMENT
# ============================================================
@user_passes_test(superadmin_required, login_url='/login/')
def users(request):
    all_users     = UserProfile.objects.select_related('user', 'branch').order_by('-id')
    branches      = Branch.objects.filter(status='Enabled')
    role_filter   = request.GET.get('role', '')
    branch_filter = request.GET.get('branch', '')
    if role_filter:
        all_users = all_users.filter(role=role_filter)
    if branch_filter:
        all_users = all_users.filter(branch_id=branch_filter)
    return render(request, 'dash/users/users.html', {
        'users': all_users,
        'branches': branches,
        'role_filter': role_filter,
        'branch_filter': branch_filter,
    })


@user_passes_test(superadmin_required, login_url='/login/')
def add_user(request):
    branches     = Branch.objects.filter(status='Enabled')
    all_profiles = UserProfile.objects.select_related('user').filter(status='Enabled')
    if request.method == 'POST':
        first_name    = request.POST.get('first_name')
        last_name     = request.POST.get('last_name')
        username      = request.POST.get('username')
        email         = request.POST.get('email')
        password      = request.POST.get('password')
        role          = request.POST.get('role')
        branch_id     = request.POST.get('branch')
        phone         = request.POST.get('phone')
        reports_to_id = request.POST.get('reports_to')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
        else:
            user = User.objects.create_user(
                username=username, email=email,
                password=password, first_name=first_name, last_name=last_name
            )
            profile = UserProfile(
                user=user, role=role, phone=phone,
                branch_id=branch_id if branch_id else None,
                reports_to_id=reports_to_id if reports_to_id else None,
            )
            if request.FILES.get('profile_pic'):
                profile.profile_pic = request.FILES['profile_pic']
            profile.save()
            messages.success(request, 'User created successfully.')
            return redirect('users')
    return render(request, 'dash/users/add_user.html', {
        'branches': branches,
        'all_profiles': all_profiles,
    })


@user_passes_test(superadmin_required, login_url='/login/')
def edit_user(request, id):
    profile      = get_object_or_404(UserProfile, id=id)
    branches     = Branch.objects.filter(status='Enabled')
    all_profiles = UserProfile.objects.select_related('user').filter(status='Enabled').exclude(id=id)
    if request.method == 'POST':
        profile.user.first_name = request.POST.get('first_name')
        profile.user.last_name  = request.POST.get('last_name')
        profile.user.email      = request.POST.get('email')
        profile.user.save()
        profile.role          = request.POST.get('role')
        profile.phone         = request.POST.get('phone')
        profile.branch_id     = request.POST.get('branch') or None
        profile.reports_to_id = request.POST.get('reports_to') or None
        if request.FILES.get('profile_pic'):
            profile.profile_pic = request.FILES['profile_pic']
        profile.save()
        messages.success(request, 'User updated successfully.')
        return redirect('users')
    return render(request, 'dash/users/edit_user.html', {
        'data': profile,
        'branches': branches,
        'all_profiles': all_profiles,
    })


@user_passes_test(superadmin_required, login_url='/login/')
def delete_user(request, id):
    profile = get_object_or_404(UserProfile, id=id)
    profile.user.delete()
    messages.success(request, 'User deleted.')
    return redirect('users')


@user_passes_test(superadmin_required, login_url='/login/')
def enable_user(request, id):
    item = get_object_or_404(UserProfile, id=id)
    item.status = 'Enabled'
    item.save()
    return redirect('users')


@user_passes_test(superadmin_required, login_url='/login/')
def disable_user(request, id):
    item = get_object_or_404(UserProfile, id=id)
    item.status = 'Disabled'
    item.save()
    return redirect('users')


# ============================================================
# LEAD MANAGEMENT
# ============================================================
@user_passes_test(superadmin_required, login_url='/login/')
def leads(request):
    all_leads     = Lead.objects.select_related('branch', 'assigned_to__user').order_by('-created_at')
    branches      = Branch.objects.filter(status='Enabled')
    stage_filter  = request.GET.get('stage', '')
    temp_filter   = request.GET.get('temperature', '')
    source_filter = request.GET.get('source', '')
    branch_filter = request.GET.get('branch', '')
    search        = request.GET.get('q', '')
    if stage_filter:
        all_leads = all_leads.filter(stage=stage_filter)
    if temp_filter:
        all_leads = all_leads.filter(temperature=temp_filter)
    if source_filter:
        all_leads = all_leads.filter(source=source_filter)
    if branch_filter:
        all_leads = all_leads.filter(branch_id=branch_filter)
    if search:
        all_leads = all_leads.filter(
            Q(name__icontains=search) | Q(phone__icontains=search) | Q(email__icontains=search)
        )
    return render(request, 'dash/leads/leads.html', {
        'leads': all_leads,
        'branches': branches,
        'stage_filter': stage_filter,
        'temp_filter': temp_filter,
        'source_filter': source_filter,
        'branch_filter': branch_filter,
        'search': search,
    })


@user_passes_test(superadmin_required, login_url='/login/')
def add_lead(request):
    branches = Branch.objects.filter(status='Enabled')
    members  = UserProfile.objects.filter(role='member', status='Enabled').select_related('user')
    if request.method == 'POST':
        Lead.objects.create(
            name              = request.POST.get('name'),
            phone             = request.POST.get('phone'),
            email             = request.POST.get('email') or None,
            location          = request.POST.get('location'),
            source            = request.POST.get('source'),
            campaign_name     = request.POST.get('campaign_name'),
            ad_set            = request.POST.get('ad_set'),
            ad_creative       = request.POST.get('ad_creative'),
            landing_page_url  = request.POST.get('landing_page_url') or None,
            property_type     = request.POST.get('property_type') or None,
            budget_min        = request.POST.get('budget_min') or None,
            budget_max        = request.POST.get('budget_max') or None,
            property_location = request.POST.get('property_location'),
            bhk_preference    = request.POST.get('bhk_preference') or None,
            purpose           = request.POST.get('purpose') or None,
            timeline          = request.POST.get('timeline') or None,
            readiness         = request.POST.get('readiness') or None,
            temperature       = request.POST.get('temperature', 'cold'),
            stage             = request.POST.get('stage', 'new'),
            assigned_to_id    = request.POST.get('assigned_to') or None,
            branch_id         = request.POST.get('branch') or None,
            notes             = request.POST.get('notes'),
        )
        messages.success(request, 'Lead added successfully.')
        return redirect('leads')
    return render(request, 'dash/leads/add_lead.html', {'branches': branches, 'members': members})


@user_passes_test(superadmin_required, login_url='/login/')
def edit_lead(request, id):
    item     = get_object_or_404(Lead, id=id)
    branches = Branch.objects.filter(status='Enabled')
    members  = UserProfile.objects.filter(role='member', status='Enabled').select_related('user')
    if request.method == 'POST':
        item.name              = request.POST.get('name')
        item.phone             = request.POST.get('phone')
        item.email             = request.POST.get('email') or None
        item.location          = request.POST.get('location')
        item.source            = request.POST.get('source')
        item.campaign_name     = request.POST.get('campaign_name')
        item.ad_set            = request.POST.get('ad_set')
        item.ad_creative       = request.POST.get('ad_creative')
        item.landing_page_url  = request.POST.get('landing_page_url') or None
        item.property_type     = request.POST.get('property_type') or None
        item.budget_min        = request.POST.get('budget_min') or None
        item.budget_max        = request.POST.get('budget_max') or None
        item.property_location = request.POST.get('property_location')
        item.bhk_preference    = request.POST.get('bhk_preference') or None
        item.purpose           = request.POST.get('purpose') or None
        item.timeline          = request.POST.get('timeline') or None
        item.readiness         = request.POST.get('readiness') or None
        item.temperature       = request.POST.get('temperature', 'cold')
        item.stage             = request.POST.get('stage', 'new')
        item.assigned_to_id    = request.POST.get('assigned_to') or None
        item.branch_id         = request.POST.get('branch') or None
        item.notes             = request.POST.get('notes')
        item.save()
        messages.success(request, 'Lead updated successfully.')
        return redirect('leads')
    return render(request, 'dash/leads/edit_lead.html', {
        'data': item, 'branches': branches, 'members': members
    })


@user_passes_test(superadmin_required, login_url='/login/')
def delete_lead(request, id):
    item = get_object_or_404(Lead, id=id)
    item.delete()
    messages.success(request, 'Lead deleted.')
    return redirect('leads')


@user_passes_test(superadmin_required, login_url='/login/')
def enable_lead(request, id):
    item = get_object_or_404(Lead, id=id)
    item.status = 'Enabled'
    item.save()
    return redirect('leads')


@user_passes_test(superadmin_required, login_url='/login/')
def disable_lead(request, id):
    item = get_object_or_404(Lead, id=id)
    item.status = 'Disabled'
    item.save()
    return redirect('leads')


@user_passes_test(superadmin_required, login_url='/login/')
def view_lead(request, id):
    item      = get_object_or_404(Lead, id=id)
    calls     = item.calls.order_by('-created_at')
    followups = item.followups.order_by('followup_at')
    wrapups   = item.wrapups.order_by('-created_at')
    return render(request, 'dash/leads/view_lead.html', {
        'lead': item, 'calls': calls, 'followups': followups, 'wrapups': wrapups
    })


# ============================================================
# CALL LOG MANAGEMENT
# ============================================================
@user_passes_test(superadmin_required, login_url='/login/')
def calls(request):
    all_calls      = CallLog.objects.select_related('lead', 'called_by__user', 'branch').order_by('-created_at')
    branches       = Branch.objects.filter(status='Enabled')
    branch_filter  = request.GET.get('branch', '')
    outcome_filter = request.GET.get('outcome', '')
    search         = request.GET.get('q', '')
    if branch_filter:
        all_calls = all_calls.filter(branch_id=branch_filter)
    if outcome_filter:
        all_calls = all_calls.filter(call_outcome=outcome_filter)
    if search:
        all_calls = all_calls.filter(lead__name__icontains=search)
    return render(request, 'dash/calls/calls.html', {
        'calls': all_calls, 'branches': branches,
        'branch_filter': branch_filter, 'outcome_filter': outcome_filter, 'search': search,
    })


@user_passes_test(superadmin_required, login_url='/login/')
def add_call(request):
    all_leads = Lead.objects.filter(status='Enabled').order_by('name')
    members   = UserProfile.objects.filter(role='member', status='Enabled').select_related('user')
    branches  = Branch.objects.filter(status='Enabled')
    if request.method == 'POST':
        CallLog.objects.create(
            lead_id          = request.POST.get('lead'),
            call_type        = request.POST.get('call_type'),
            call_duration    = request.POST.get('call_duration') or 0,
            call_outcome     = request.POST.get('call_outcome'),
            call_notes       = request.POST.get('call_notes'),
            next_followup_at = request.POST.get('next_followup_at') or None,
            called_by_id     = request.POST.get('called_by') or None,
            branch_id        = request.POST.get('branch') or None,
        )
        messages.success(request, 'Call log added.')
        return redirect('calls')
    return render(request, 'dash/calls/add_call.html', {
        'all_leads': all_leads, 'members': members, 'branches': branches
    })


@user_passes_test(superadmin_required, login_url='/login/')
def edit_call(request, id):
    item      = get_object_or_404(CallLog, id=id)
    all_leads = Lead.objects.filter(status='Enabled').order_by('name')
    members   = UserProfile.objects.filter(role='member', status='Enabled').select_related('user')
    branches  = Branch.objects.filter(status='Enabled')
    if request.method == 'POST':
        item.lead_id          = request.POST.get('lead')
        item.call_type        = request.POST.get('call_type')
        item.call_duration    = request.POST.get('call_duration') or 0
        item.call_outcome     = request.POST.get('call_outcome')
        item.call_notes       = request.POST.get('call_notes')
        item.next_followup_at = request.POST.get('next_followup_at') or None
        item.called_by_id     = request.POST.get('called_by') or None
        item.branch_id        = request.POST.get('branch') or None
        item.save()
        messages.success(request, 'Call log updated.')
        return redirect('calls')
    return render(request, 'dash/calls/edit_call.html', {
        'data': item, 'all_leads': all_leads, 'members': members, 'branches': branches
    })


@user_passes_test(superadmin_required, login_url='/login/')
def delete_call(request, id):
    item = get_object_or_404(CallLog, id=id)
    item.delete()
    messages.success(request, 'Call log deleted.')
    return redirect('calls')


@user_passes_test(superadmin_required, login_url='/login/')
def enable_call(request, id):
    item = get_object_or_404(CallLog, id=id)
    item.status = 'Enabled'
    item.save()
    return redirect('calls')


@user_passes_test(superadmin_required, login_url='/login/')
def disable_call(request, id):
    item = get_object_or_404(CallLog, id=id)
    item.status = 'Disabled'
    item.save()
    return redirect('calls')


# ============================================================
# FOLLOW-UP MANAGEMENT
# ============================================================
@user_passes_test(superadmin_required, login_url='/login/')
def followups(request):
    all_followups = FollowUp.objects.select_related('lead', 'assigned_to__user', 'branch').order_by('followup_at')
    branches      = Branch.objects.filter(status='Enabled')
    status_filter = request.GET.get('status', '')
    branch_filter = request.GET.get('branch', '')
    search        = request.GET.get('q', '')
    if status_filter:
        all_followups = all_followups.filter(followup_status=status_filter)
    if branch_filter:
        all_followups = all_followups.filter(branch_id=branch_filter)
    if search:
        all_followups = all_followups.filter(lead__name__icontains=search)
    return render(request, 'dash/followups/followups.html', {
        'followups': all_followups, 'branches': branches,
        'status_filter': status_filter, 'branch_filter': branch_filter, 'search': search,
    })


@user_passes_test(superadmin_required, login_url='/login/')
def add_followup(request):
    all_leads = Lead.objects.filter(status='Enabled').order_by('name')
    members   = UserProfile.objects.filter(role='member', status='Enabled').select_related('user')
    branches  = Branch.objects.filter(status='Enabled')
    if request.method == 'POST':
        FollowUp.objects.create(
            lead_id         = request.POST.get('lead'),
            followup_at     = request.POST.get('followup_at'),
            followup_type   = request.POST.get('followup_type'),
            notes           = request.POST.get('notes'),
            followup_status = request.POST.get('followup_status', 'pending'),
            assigned_to_id  = request.POST.get('assigned_to') or None,
            branch_id       = request.POST.get('branch') or None,
        )
        messages.success(request, 'Follow-up added.')
        return redirect('followups')
    return render(request, 'dash/followups/add_followup.html', {
        'all_leads': all_leads, 'members': members, 'branches': branches
    })


@user_passes_test(superadmin_required, login_url='/login/')
def edit_followup(request, id):
    item      = get_object_or_404(FollowUp, id=id)
    all_leads = Lead.objects.filter(status='Enabled').order_by('name')
    members   = UserProfile.objects.filter(role='member', status='Enabled').select_related('user')
    branches  = Branch.objects.filter(status='Enabled')
    if request.method == 'POST':
        item.lead_id         = request.POST.get('lead')
        item.followup_at     = request.POST.get('followup_at')
        item.followup_type   = request.POST.get('followup_type')
        item.notes           = request.POST.get('notes')
        item.followup_status = request.POST.get('followup_status', 'pending')
        item.assigned_to_id  = request.POST.get('assigned_to') or None
        item.branch_id       = request.POST.get('branch') or None
        item.save()
        messages.success(request, 'Follow-up updated.')
        return redirect('followups')
    return render(request, 'dash/followups/edit_followup.html', {
        'data': item, 'all_leads': all_leads, 'members': members, 'branches': branches
    })


@user_passes_test(superadmin_required, login_url='/login/')
def delete_followup(request, id):
    item = get_object_or_404(FollowUp, id=id)
    item.delete()
    messages.success(request, 'Follow-up deleted.')
    return redirect('followups')


@user_passes_test(superadmin_required, login_url='/login/')
def enable_followup(request, id):
    item = get_object_or_404(FollowUp, id=id)
    item.status = 'Enabled'
    item.save()
    return redirect('followups')


@user_passes_test(superadmin_required, login_url='/login/')
def disable_followup(request, id):
    item = get_object_or_404(FollowUp, id=id)
    item.status = 'Disabled'
    item.save()
    return redirect('followups')


# ============================================================
# WRAP-UP MANAGEMENT
# ============================================================
@user_passes_test(superadmin_required, login_url='/login/')
def wrapups(request):
    all_wrapups   = CallWrapUp.objects.select_related('lead', 'submitted_by__user').order_by('-created_at')
    branch_filter = request.GET.get('branch', '')
    locked_filter = request.GET.get('locked', '')
    if branch_filter:
        all_wrapups = all_wrapups.filter(lead__branch_id=branch_filter)
    if locked_filter == '1':
        all_wrapups = all_wrapups.filter(is_locked=True)
    elif locked_filter == '0':
        all_wrapups = all_wrapups.filter(is_locked=False)
    branches = Branch.objects.filter(status='Enabled')
    return render(request, 'dash/calls/wrapups.html', {
        'wrapups': all_wrapups, 'branches': branches,
        'branch_filter': branch_filter, 'locked_filter': locked_filter,
    })


# ============================================================
# SETTINGS
# ============================================================
@user_passes_test(superadmin_required, login_url='/login/')
def settings_view(request):
    setting_keys  = ['wrapup_window_minutes', 'lead_auto_assign', 'whatsapp_provider']
    settings_data = {}
    for key in setting_keys:
        obj, _ = SystemSetting.objects.get_or_create(key=key, defaults={'value': ''})
        settings_data[key] = obj
    if request.method == 'POST':
        for key in setting_keys:
            val = request.POST.get(key, '')
            SystemSetting.objects.filter(key=key).update(value=val)
        messages.success(request, 'Settings saved.')
        return redirect('settings')
    return render(request, 'dash/settings/settings.html', {'settings': settings_data})


# ============================================================
# BULK LEAD UPLOAD
# ============================================================
@user_passes_test(superadmin_required, login_url='/login/')
def bulk_upload_leads(request):
    branches = Branch.objects.filter(status='Enabled')
    members  = UserProfile.objects.filter(role='member', status='Enabled').select_related('user')

    if request.method == 'POST':
        uploaded_file  = request.FILES.get('bulk_file')
        branch_id      = request.POST.get('branch') or None
        assigned_to_id = request.POST.get('assigned_to') or None

        if not uploaded_file:
            messages.error(request, 'Please select a file to upload.')
            return redirect('bulk_upload_leads')

        filename = uploaded_file.name.lower()
        rows     = []
        errors   = []

        try:
            if filename.endswith('.csv'):
                decoded = uploaded_file.read().decode('utf-8-sig')
                reader  = csv.DictReader(io.StringIO(decoded))
                for i, row in enumerate(reader, start=2):
                    rows.append((i, row))

            elif filename.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
                ws = wb.active
                headers = [
                    str(cell.value).strip().lower().replace(' ', '_') if cell.value else ''
                    for cell in next(ws.iter_rows(min_row=1, max_row=1))
                ]
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    row_dict = {
                        headers[j]: (str(val).strip() if val is not None else '')
                        for j, val in enumerate(row)
                    }
                    rows.append((i, row_dict))
                wb.close()
            else:
                messages.error(request, 'Unsupported file. Upload .csv or .xlsx only.')
                return redirect('bulk_upload_leads')

        except Exception as e:
            messages.error(request, f'Error reading file: {str(e)}')
            return redirect('bulk_upload_leads')

        VALID_SOURCES = ['meta','google','website','social','campaign','referral','walkin','whatsapp','other']
        VALID_TEMPS   = ['hot','warm','cold']
        VALID_STAGES  = ['new','contacted','interested','site_visit','negotiation','closed','lost']

        created_count = 0
        skip_count    = 0

        for row_num, row in rows:
            if not any(row.values()):
                skip_count += 1
                continue

            name  = row.get('name', '').strip()
            phone = row.get('phone', '').strip()

            if not name or not phone:
                errors.append(f'Row {row_num}: Skipped — name and phone are required.')
                skip_count += 1
                continue

            source = row.get('source', 'other').strip().lower()
            if source not in VALID_SOURCES:
                source = 'other'

            temperature = row.get('temperature', 'cold').strip().lower()
            if temperature not in VALID_TEMPS:
                temperature = 'cold'

            stage = row.get('stage', 'new').strip().lower().replace(' ', '_')
            if stage not in VALID_STAGES:
                stage = 'new'

            def safe_decimal(val):
                try:
                    return float(str(val).replace(',', '').strip()) if val else None
                except:
                    return None

            try:
                Lead.objects.create(
                    name              = name,
                    phone             = phone,
                    email             = row.get('email', '') or None,
                    location          = row.get('location', ''),
                    source            = source,
                    campaign_name     = row.get('campaign_name', ''),
                    ad_set            = row.get('ad_set', ''),
                    ad_creative       = row.get('ad_creative', ''),
                    landing_page_url  = row.get('landing_page_url', '') or None,
                    property_type     = row.get('property_type', '') or None,
                    budget_min        = safe_decimal(row.get('budget_min')),
                    budget_max        = safe_decimal(row.get('budget_max')),
                    property_location = row.get('property_location', ''),
                    bhk_preference    = row.get('bhk_preference', '') or None,
                    purpose           = row.get('purpose', '') or None,
                    timeline          = row.get('timeline', '') or None,
                    readiness         = row.get('readiness', '') or None,
                    temperature       = temperature,
                    stage             = stage,
                    notes             = row.get('notes', ''),
                    branch_id         = branch_id,
                    assigned_to_id    = assigned_to_id,
                )
                created_count += 1
            except Exception as e:
                errors.append(f'Row {row_num}: Error — {str(e)}')
                skip_count += 1

        if created_count:
            messages.success(request, f'{created_count} leads imported successfully.')
        if errors:
            for err in errors[:10]:
                messages.warning(request, err)
        if skip_count and not errors:
            messages.warning(request, f'{skip_count} rows skipped.')

        return redirect('leads')

    return render(request, 'dash/leads/bulk_upload.html', {
        'branches': branches,
        'members': members,
    })


@user_passes_test(superadmin_required, login_url='/login/')
def download_lead_template(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="CoreCRM_Lead_Template.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'name', 'phone', 'email', 'location',
        'source', 'campaign_name', 'ad_set', 'ad_creative', 'landing_page_url',
        'property_type', 'budget_min', 'budget_max', 'property_location',
        'bhk_preference', 'purpose', 'timeline', 'readiness',
        'temperature', 'stage', 'notes'
    ])
    writer.writerow([
        'Rahul Sharma', '9876543210', 'rahul@email.com', 'Salt Lake Kolkata',
        'meta', 'Diwali Campaign 2026', 'Ad Set 1', 'Creative A', '',
        'apartment', '3000000', '6000000', 'Newtown',
        '2bhk', 'selfuse', '3months', 'ready',
        'warm', 'new', 'Interested in 2BHK near IT hub'
    ])
    return response


# ============================================================
# HR PANEL
# ============================================================
from .models import Attendance, LeaveRequest, Payroll

@user_passes_test(superadmin_required, login_url='/login/')
def hr_panel(request):
    branches  = Branch.objects.filter(status='Enabled')
    employees = UserProfile.objects.filter(status='Enabled').select_related('user', 'branch')
    branch_filter = request.GET.get('branch', '')
    role_filter   = request.GET.get('role', '')
    search        = request.GET.get('q', '')
    if branch_filter:
        employees = employees.filter(branch_id=branch_filter)
    if role_filter:
        employees = employees.filter(role=role_filter)
    if search:
        employees = employees.filter(
            Q(user__first_name__icontains=search) | Q(user__last_name__icontains=search)
        )
    total_employees = employees.count()
    present_today   = Attendance.objects.filter(date=timezone.now().date(), status='present').count()
    on_leave_today  = Attendance.objects.filter(date=timezone.now().date(), status='on_leave').count()
    pending_leaves  = LeaveRequest.objects.filter(leave_status='pending').count()
    return render(request, 'dash/hr/hr_panel.html', {
        'employees': employees,
        'branches': branches,
        'branch_filter': branch_filter,
        'role_filter': role_filter,
        'search': search,
        'total_employees': total_employees,
        'present_today': present_today,
        'on_leave_today': on_leave_today,
        'pending_leaves': pending_leaves,
    })


@user_passes_test(superadmin_required, login_url='/login/')
def attendance(request):
    branches  = Branch.objects.filter(status='Enabled')
    records   = Attendance.objects.select_related('employee__user', 'branch').order_by('-date', '-created_at')
    branch_filter = request.GET.get('branch', '')
    date_filter   = request.GET.get('date', '')
    status_filter = request.GET.get('status', '')
    search        = request.GET.get('q', '')
    if branch_filter:
        records = records.filter(branch_id=branch_filter)
    if date_filter:
        records = records.filter(date=date_filter)
    if status_filter:
        records = records.filter(status=status_filter)
    if search:
        records = records.filter(
            Q(employee__user__first_name__icontains=search) |
            Q(employee__user__last_name__icontains=search)
        )
    return render(request, 'dash/hr/attendance.html', {
        'records': records, 'branches': branches,
        'branch_filter': branch_filter, 'date_filter': date_filter,
        'status_filter': status_filter, 'search': search,
    })


@user_passes_test(superadmin_required, login_url='/login/')
def add_attendance(request):
    employees = UserProfile.objects.filter(status='Enabled').select_related('user')
    branches  = Branch.objects.filter(status='Enabled')
    if request.method == 'POST':
        record = Attendance.objects.create(
            employee_id    = request.POST.get('employee'),
            branch_id      = request.POST.get('branch') or None,
            date           = request.POST.get('date'),
            punch_in       = request.POST.get('punch_in') or None,
            punch_out      = request.POST.get('punch_out') or None,
            status         = request.POST.get('status', 'present'),
            is_out_of_zone = request.POST.get('is_out_of_zone') == 'on',
            notes          = request.POST.get('notes', ''),
        )
        record.calculate_hours()
        messages.success(request, 'Attendance record added.')
        return redirect('attendance')
    return render(request, 'dash/hr/add_attendance.html', {
        'employees': employees, 'branches': branches
    })


@user_passes_test(superadmin_required, login_url='/login/')
def edit_attendance(request, id):
    record    = get_object_or_404(Attendance, id=id)
    employees = UserProfile.objects.filter(status='Enabled').select_related('user')
    branches  = Branch.objects.filter(status='Enabled')
    if request.method == 'POST':
        record.employee_id    = request.POST.get('employee')
        record.branch_id      = request.POST.get('branch') or None
        record.date           = request.POST.get('date')
        record.punch_in       = request.POST.get('punch_in') or None
        record.punch_out      = request.POST.get('punch_out') or None
        record.status         = request.POST.get('status', 'present')
        record.is_out_of_zone = request.POST.get('is_out_of_zone') == 'on'
        record.notes          = request.POST.get('notes', '')
        record.save()
        record.calculate_hours()
        messages.success(request, 'Attendance updated.')
        return redirect('attendance')
    return render(request, 'dash/hr/edit_attendance.html', {
        'data': record, 'employees': employees, 'branches': branches
    })


@user_passes_test(superadmin_required, login_url='/login/')
def delete_attendance(request, id):
    get_object_or_404(Attendance, id=id).delete()
    messages.success(request, 'Record deleted.')
    return redirect('attendance')


@user_passes_test(superadmin_required, login_url='/login/')
def leaves(request):
    branches  = Branch.objects.filter(status='Enabled')
    records   = LeaveRequest.objects.select_related('employee__user', 'approved_by__user').order_by('-created_at')
    status_filter = request.GET.get('status', '')
    branch_filter = request.GET.get('branch', '')
    search        = request.GET.get('q', '')
    if status_filter:
        records = records.filter(leave_status=status_filter)
    if branch_filter:
        records = records.filter(employee__branch_id=branch_filter)
    if search:
        records = records.filter(
            Q(employee__user__first_name__icontains=search) |
            Q(employee__user__last_name__icontains=search)
        )
    return render(request, 'dash/hr/leaves.html', {
        'records': records, 'branches': branches,
        'status_filter': status_filter, 'branch_filter': branch_filter, 'search': search,
    })


@user_passes_test(superadmin_required, login_url='/login/')
def add_leave(request):
    employees = UserProfile.objects.filter(status='Enabled').select_related('user')
    if request.method == 'POST':
        LeaveRequest.objects.create(
            employee_id  = request.POST.get('employee'),
            leave_type   = request.POST.get('leave_type'),
            from_date    = request.POST.get('from_date'),
            to_date      = request.POST.get('to_date'),
            reason       = request.POST.get('reason'),
            leave_status = request.POST.get('leave_status', 'pending'),
        )
        messages.success(request, 'Leave request added.')
        return redirect('leaves')
    return render(request, 'dash/hr/add_leave.html', {'employees': employees})


@user_passes_test(superadmin_required, login_url='/login/')
def edit_leave(request, id):
    record    = get_object_or_404(LeaveRequest, id=id)
    employees = UserProfile.objects.filter(status='Enabled').select_related('user')
    profiles  = UserProfile.objects.filter(status='Enabled').select_related('user')
    if request.method == 'POST':
        record.employee_id    = request.POST.get('employee')
        record.leave_type     = request.POST.get('leave_type')
        record.from_date      = request.POST.get('from_date')
        record.to_date        = request.POST.get('to_date')
        record.reason         = request.POST.get('reason')
        record.leave_status   = request.POST.get('leave_status', 'pending')
        record.approved_by_id = request.POST.get('approved_by') or None
        record.remarks        = request.POST.get('remarks', '')
        record.save()
        messages.success(request, 'Leave updated.')
        return redirect('leaves')
    return render(request, 'dash/hr/edit_leave.html', {
        'data': record, 'employees': employees, 'profiles': profiles
    })


@user_passes_test(superadmin_required, login_url='/login/')
def delete_leave(request, id):
    get_object_or_404(LeaveRequest, id=id).delete()
    messages.success(request, 'Leave request deleted.')
    return redirect('leaves')


@user_passes_test(superadmin_required, login_url='/login/')
def approve_leave(request, id):
    record = get_object_or_404(LeaveRequest, id=id)
    record.leave_status = 'approved'
    record.approved_by  = request.user.profile
    record.save()
    messages.success(request, 'Leave approved.')
    return redirect('leaves')


@user_passes_test(superadmin_required, login_url='/login/')
def reject_leave(request, id):
    record = get_object_or_404(LeaveRequest, id=id)
    record.leave_status = 'rejected'
    record.approved_by  = request.user.profile
    record.save()
    messages.success(request, 'Leave rejected.')
    return redirect('leaves')


@user_passes_test(superadmin_required, login_url='/login/')
def payroll(request):
    branches  = Branch.objects.filter(status='Enabled')
    records   = Payroll.objects.select_related('employee__user').order_by('-year', '-created_at')
    branch_filter = request.GET.get('branch', '')
    month_filter  = request.GET.get('month', '')
    search        = request.GET.get('q', '')
    if branch_filter:
        records = records.filter(employee__branch_id=branch_filter)
    if month_filter:
        records = records.filter(month=month_filter)
    if search:
        records = records.filter(
            Q(employee__user__first_name__icontains=search) |
            Q(employee__user__last_name__icontains=search)
        )
    return render(request, 'dash/hr/payroll.html', {
        'records': records, 'branches': branches,
        'branch_filter': branch_filter, 'month_filter': month_filter, 'search': search,
    })


@user_passes_test(superadmin_required, login_url='/login/')
def add_payroll(request):
    employees = UserProfile.objects.filter(status='Enabled').select_related('user')
    if request.method == 'POST':
        base       = float(request.POST.get('base_salary', 0))
        bonus      = float(request.POST.get('bonus', 0))
        deductions = float(request.POST.get('deductions', 0))
        Payroll.objects.create(
            employee_id = request.POST.get('employee'),
            month       = request.POST.get('month'),
            year        = request.POST.get('year'),
            base_salary = base,
            bonus       = bonus,
            deductions  = deductions,
            net_salary  = base + bonus - deductions,
            notes       = request.POST.get('notes', ''),
        )
        messages.success(request, 'Payroll record added.')
        return redirect('payroll')
    return render(request, 'dash/hr/add_payroll.html', {'employees': employees})


@user_passes_test(superadmin_required, login_url='/login/')
def edit_payroll(request, id):
    record    = get_object_or_404(Payroll, id=id)
    employees = UserProfile.objects.filter(status='Enabled').select_related('user')
    if request.method == 'POST':
        base       = float(request.POST.get('base_salary', 0))
        bonus      = float(request.POST.get('bonus', 0))
        deductions = float(request.POST.get('deductions', 0))
        record.employee_id = request.POST.get('employee')
        record.month       = request.POST.get('month')
        record.year        = request.POST.get('year')
        record.base_salary = base
        record.bonus       = bonus
        record.deductions  = deductions
        record.net_salary  = base + bonus - deductions
        record.notes       = request.POST.get('notes', '')
        record.save()
        messages.success(request, 'Payroll updated.')
        return redirect('payroll')
    return render(request, 'dash/hr/edit_payroll.html', {
        'data': record, 'employees': employees
    })


@user_passes_test(superadmin_required, login_url='/login/')
def delete_payroll(request, id):
    get_object_or_404(Payroll, id=id).delete()
    messages.success(request, 'Payroll record deleted.')
    return redirect('payroll')


@user_passes_test(superadmin_required, login_url='/login/')
def employee_detail(request, id):
    employee        = get_object_or_404(UserProfile, id=id)
    branches        = Branch.objects.filter(status='Enabled')
    att_records     = Attendance.objects.filter(employee=employee).order_by('-date')
    leave_records   = LeaveRequest.objects.filter(employee=employee).order_by('-created_at')
    payroll_records = Payroll.objects.filter(employee=employee).order_by('-year', '-created_at')

    if request.method == 'POST':
        record = Attendance.objects.create(
            employee       = employee,
            branch_id      = request.POST.get('branch') or employee.branch_id,
            date           = request.POST.get('date'),
            punch_in       = request.POST.get('punch_in') or None,
            punch_out      = request.POST.get('punch_out') or None,
            status         = request.POST.get('status', 'present'),
            is_out_of_zone = request.POST.get('is_out_of_zone') == 'on',
            notes          = request.POST.get('notes', ''),
        )
        record.calculate_hours()
        messages.success(request, f'Attendance added for {employee.user.get_full_name()}.')
        return redirect('employee_detail', id=id)

    return render(request, 'dash/hr/employee_detail.html', {
        'employee': employee,
        'att_records': att_records,
        'leave_records': leave_records,
        'payroll_records': payroll_records,
        'branches': branches,
    })